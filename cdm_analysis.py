"""
CDM Usage Analysis Script
==========================
Analyzes CDM (Cash Deposit Machine) transaction data to answer:
  1. Device usage (volume & value per machine) — overall AND per-device
  2. Busiest time of day (hourly load curve) — overall AND per-device
  3. When usage starts to peak (ramp-up pattern) — per-device
  4. Day-of-week patterns — overall AND per-device
  5. Declined/failed transactions — count, %, reasons, by device and by hour
     (declined transactions are NEVER dropped from the data — they're kept,
     counted, and flagged everywhere: charts, summary, and highlighted red
     in the Excel report)

INPUT: An Excel/CSV export with these columns (matches the anonymized
       bank export format this was built against):
   Code, Created, Created Time, Updated Time, Device Code, Device Name,
   Transaction Display Name, Amount, Event Display Name, Result Code,
   Result Message, Status

USAGE (local / any environment):
   python cdm_analysis.py /path/to/CDM_export.xlsx /path/to/output_folder

USAGE (Google Colab): see CDM_Automated_Analysis.ipynb
"""

import sys
import os
import gc
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

DOW_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

# ---------- Professional styling ----------
plt.rcParams.update({
    "figure.facecolor": "white",
    "axes.facecolor": "white",
    "axes.edgecolor": "#333333",
    "axes.linewidth": 0.8,
    "axes.grid": False,                 # no gridlines
    "font.family": "DejaVu Sans",
    "font.size": 11,
    "axes.titlesize": 14,
    "axes.titleweight": "bold",
    "axes.labelsize": 11,
    "axes.labelcolor": "#333333",
    "xtick.color": "#333333",
    "ytick.color": "#333333",
    "legend.frameon": False,
    "figure.dpi": 150,
    "savefig.dpi": 180,
})

# Professional, muted palette (navy / slate / amber family)
DEVICE_COLORS = ["#1e3a5f", "#c17817", "#4a7c6f", "#8b3a3a", "#6b5b95", "#2f6690"]
DECLINE_COLOR = "#b91c1c"
NEUTRAL_COLOR = "#94a3b8"
ACCENT_COLOR = "#1e3a5f"

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# ---------- Human-friendly labels ----------
# The raw export uses internal/technical names for each session step and
# numeric result codes for failures. These aren't easy for a general
# audience to read, so we translate them here. The original raw values are
# still kept in the Excel export (Result Code / Result Message / Event
# Display Name columns) for anyone who needs to trace back to source data.

STEP_LABELS = {
    "Cash Deposit": "Cash Deposit",
    "Generate Transaction Id": "Session Started",
    "TITLE_FETCH": "Account Lookup",
    "VERIFY_BIOMETRIC_VERIFICATION": "Biometric Verification",
    "Safe Watch": "Security Check",
    "Generate Token": "Token Generated",
    "Account Inquiry": "Account Balance Inquiry",
}

def friendly_step(name):
    return STEP_LABELS.get(name, name)


# Maps a Result Code to a plain-English reason. Codes not listed here fall
# back to a generic "Other Error (Code N)" label. Update this mapping if
# your bank/vendor provides an official code dictionary.
DECLINE_REASON_LABELS = {
    58: "System / Processing Error",
    1: "Transaction Not Completed",
    204: "System / Processing Error",
    401: "Transaction Not Completed",
}

def friendly_decline_reason(code):
    try:
        code_int = int(code)
    except (ValueError, TypeError):
        return f"Other Error (Code {code})"
    label = DECLINE_REASON_LABELS.get(code_int, f"Other Error (Code {code_int})")
    return f"{label} (Code {code_int})"


def finish_chart(fig, ax_or_axes, subtitle=None):
    """Apply consistent professional finishing touches to any chart:
    strip top/right spines, thin remaining spines, add a subtitle and a
    generation-date footer. Call this right before plt.savefig()."""
    axes_list = ax_or_axes if isinstance(ax_or_axes, (list, tuple)) else \
        (list(ax_or_axes) if hasattr(ax_or_axes, "__iter__") and not hasattr(ax_or_axes, "plot") else [ax_or_axes])
    for ax in axes_list:
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#cccccc")
        ax.spines["bottom"].set_color("#cccccc")
        ax.tick_params(length=3)

    if subtitle:
        fig.text(0.5, 0.955, subtitle, ha="center", fontsize=9.5, color="#666666", style="italic")

    fig.text(
        0.99, 0.01, f"Generated {pd.Timestamp.now().strftime('%d %b %Y')} · CDM Usage Analysis",
        ha="right", va="bottom", fontsize=7.5, color="#999999"
    )


def fmt_pkr(value):
    """Format a PKR amount into a compact, readable string, e.g. 304994800 -> 'PKR 305.0M'."""
    if abs(value) >= 1e6:
        return f"PKR {value/1e6:.1f}M"
    if abs(value) >= 1e3:
        return f"PKR {value/1e3:.0f}K"
    return f"PKR {value:.0f}"


def load_data(filepath):
    """Load the raw export (xlsx or csv) and build a proper datetime column."""
    if filepath.lower().endswith(".csv"):
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    required = ["Created", "Created Time", "Device Name", "Amount", "Event Display Name"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")

    df["datetime"] = pd.to_datetime(
        df["Created"].astype(str) + " " + df["Created Time"].astype(str),
        errors="coerce"
    )
    n_bad = df["datetime"].isna().sum()
    if n_bad:
        print(f"Warning: {n_bad} rows had unparseable dates and were dropped.")
        df = df.dropna(subset=["datetime"])

    df["hour"] = df["datetime"].dt.hour
    df["dow"] = df["datetime"].dt.day_name()
    df["date"] = df["datetime"].dt.date
    return df


def get_deposits(df):
    """All deposit attempts — SUCCESSFUL AND DECLINED/INCOMPLETE.
    Nothing is dropped here. We use the 'Cash Deposit' event rows as the
    true usage unit, rather than every intermediate session step (token
    generation, biometric verification, etc.), since those are sub-steps
    of a single visit. Each row keeps its real Status so declines can be
    tracked and reported rather than discarded."""
    deposits = df[df["Event Display Name"] == "Cash Deposit"].copy()
    deposits["declined"] = deposits["Status"].astype(str).str.upper() != "SUCCESS"

    def reason(row):
        if not row["declined"]:
            return ""
        return friendly_decline_reason(row.get("Result Code", ""))

    deposits["decline_reason"] = deposits.apply(reason, axis=1)
    return deposits


def _hourly_series(data):
    return data.groupby("hour").size().reindex(range(24), fill_value=0)


def _device_report(d, sub, colors, outdir, idx):
    """Generate individual charts + return summary lines for ONE device."""
    lines = [f"----- {d} -----"]
    total = len(sub)
    declined = int(sub["declined"].sum())
    successful = total - declined
    decline_pct = (declined / total * 100) if total else 0
    lines.append(
        f"Transactions: {total} | Successful: {successful} | Declined: {declined} "
        f"({decline_pct:.2f}%) | Total value (successful): PKR {sub[~sub['declined']]['Amount'].sum():,.0f}"
    )

    hourly_all = _hourly_series(sub)
    hourly_declined = _hourly_series(sub[sub["declined"]])
    peak_hour = hourly_all.idxmax()
    peak_val = hourly_all.max()
    threshold = 0.5 * peak_val if peak_val else 0
    ramp_hour = next((h for h in range(24) if hourly_all[h] >= threshold), None)
    active_hours = hourly_all[hourly_all > 0].index.tolist()
    lines.append(
        f"Peak hour: {peak_hour}:00 ({peak_val} deposits) | "
        f"Ramps to 50%+ of peak by ~{ramp_hour}:00 | "
        f"Active hours: {min(active_hours) if active_hours else '-'}:00-"
        f"{max(active_hours) if active_hours else '-'}:00"
    )

    dow_counts = sub["dow"].value_counts().reindex(DOW_ORDER, fill_value=0)
    busiest_day = dow_counts.idxmax()
    lines.append(f"Busiest day: {busiest_day} ({dow_counts.max()} deposits)")

    if declined:
        reasons = sub[sub["declined"]]["decline_reason"].value_counts()
        lines.append("Decline reasons: " + "; ".join(f"{r} x{c}" for r, c in reasons.items()))
    lines.append("")

    # Chart: hourly pattern with declines overlaid
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(hourly_all.index, hourly_all.values, color=colors.get(d, ACCENT_COLOR), label="Total attempts")
    ax.bar(hourly_declined.index, hourly_declined.values, color=DECLINE_COLOR, label="Declined")
    ax.set_xticks(range(24))
    ax.set_xlabel("Hour of Day (24h)")
    ax.set_ylabel("Number of Deposits")
    ax.set_title(f"{d} — Hourly Usage Pattern")
    ax.legend()
    finish_chart(fig, ax, subtitle="Declined transactions highlighted in red")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    safe_name = "".join(c if c.isalnum() else "_" for c in d)
    fig_path = os.path.join(outdir, f"device_{idx:02d}_{safe_name}_hourly.png")
    plt.savefig(fig_path)
    plt.close()
    gc.collect()

    # Chart: day-of-week for this device
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(dow_counts.index, dow_counts.values, color=colors.get(d, ACCENT_COLOR))
    ax.set_title(f"{d} — Transactions by Day of Week")
    ax.set_ylabel("Number of Deposits")
    ax.tick_params(axis='x', rotation=30)
    finish_chart(fig, ax)
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.96, wspace=0.3, hspace=0.45)
    fig_path2 = os.path.join(outdir, f"device_{idx:02d}_{safe_name}_dow.png")
    plt.savefig(fig_path2)
    plt.close()
    gc.collect()

    return lines, {
        "total": total, "successful": successful, "declined": declined,
        "decline_pct": decline_pct, "peak_hour": peak_hour, "busiest_day": busiest_day,
    }


def analyze(df, outdir):
    os.makedirs(outdir, exist_ok=True)
    deposits = get_deposits(df)
    devices = sorted(deposits["Device Name"].unique())
    colors = {d: DEVICE_COLORS[i % len(DEVICE_COLORS)] for i, d in enumerate(devices)}

    summary_lines = []

    # ---------- 0. DATA OVERVIEW (explains the full row count vs deposit count) ----------
    total_raw_rows = len(df)
    event_breakdown = df["Event Display Name"].value_counts()
    event_breakdown_friendly = event_breakdown.rename(index=friendly_step)
    n_cash_deposit = int(event_breakdown.get("Cash Deposit", 0))
    n_other_steps = total_raw_rows - n_cash_deposit

    summary_lines.append("=== DATA OVERVIEW ===")
    summary_lines.append(
        f"Total rows in raw file: {total_raw_rows}\n"
        f"  - 'Cash Deposit' rows (actual deposit attempts): {n_cash_deposit}\n"
        f"  - Other session-step rows (not separate deposits, e.g. session start, "
        f"account lookup, biometric verification): {n_other_steps}\n"
        "Only 'Cash Deposit' rows are counted as usage below, since the other rows are "
        "sub-steps of the same customer visit, not separate transactions."
    )
    summary_lines.append("\nFull breakdown by step type:")
    summary_lines.append(event_breakdown_friendly.to_string())
    summary_lines.append("")

    fig, ax = plt.subplots(figsize=(10, 5))
    ordered = event_breakdown_friendly.sort_values(ascending=True)
    bar_colors = [ACCENT_COLOR if name == "Cash Deposit" else NEUTRAL_COLOR for name in ordered.index]
    bars = ax.barh(ordered.index, ordered.values, color=bar_colors)
    ax.bar_label(bars, padding=3)
    ax.set_title("All Rows in Raw File by Step Type")
    ax.set_xlabel("Row Count")
    finish_chart(fig, ax, subtitle="Highlighted bar = actual deposits counted in this analysis")
    plt.subplots_adjust(left=0.22, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "00_data_overview_row_breakdown.png"))
    plt.close()
    gc.collect()

    # ---------- 1. DEVICE USAGE SUMMARY (successful transactions for value) ----------
    successful = deposits[~deposits["declined"]]
    usage = successful.groupby("Device Name").agg(
        transactions=("Amount", "count"),
        total_amount=("Amount", "sum"),
        avg_amount=("Amount", "mean"),
    ).sort_values("transactions", ascending=False)
    usage["share_of_transactions_%"] = (usage["transactions"] / usage["transactions"].sum() * 100).round(1)
    usage["share_of_value_%"] = (usage["total_amount"] / usage["total_amount"].sum() * 100).round(1)
    usage.to_csv(os.path.join(outdir, "device_usage_summary.csv"))
    summary_lines.append("=== DEVICE USAGE SUMMARY (successful transactions only) ===")
    summary_lines.append(usage.to_string())
    summary_lines.append("")

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    bars = axes[0].bar(usage.index, usage["transactions"], color=[colors[d] for d in usage.index])
    axes[0].set_title("Successful Transaction Volume by Device")
    axes[0].set_ylabel("Number of Deposits")
    axes[0].bar_label(bars, padding=3)
    axes[0].tick_params(axis='x', rotation=15)

    bars2 = axes[1].bar(usage.index, usage["total_amount"], color=[colors[d] for d in usage.index])
    axes[1].set_title("Total Value Deposited by Device")
    axes[1].set_ylabel("PKR")
    axes[1].bar_label(bars2, labels=[fmt_pkr(v) for v in usage["total_amount"]], padding=3)
    axes[1].tick_params(axis='x', rotation=15)
    finish_chart(fig, axes, subtitle="Successful transactions only (declined excluded from value totals)")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "01_device_usage.png"))
    plt.close()
    gc.collect()

    # ---------- 2. DECLINED TRANSACTIONS — OVERALL ----------
    total_all = len(deposits)
    total_declined = int(deposits["declined"].sum())
    decline_pct_overall = (total_declined / total_all * 100) if total_all else 0
    summary_lines.append("=== DECLINED / FAILED TRANSACTIONS (OVERALL) ===")
    summary_lines.append(
        f"Total attempts: {total_all} | Successful: {total_all - total_declined} | "
        f"Declined: {total_declined} ({decline_pct_overall:.2f}%)"
    )
    decline_by_device = deposits.groupby("Device Name")["declined"].agg(
        declined_count="sum", total_count="count"
    )
    decline_by_device["decline_pct"] = (decline_by_device["declined_count"] / decline_by_device["total_count"] * 100).round(2)
    summary_lines.append("\nDecline rate by device:")
    summary_lines.append(decline_by_device.to_string())

    reason_counts = deposits[deposits["declined"]]["decline_reason"].value_counts()
    summary_lines.append("\nDecline reasons (all devices):")
    summary_lines.append(reason_counts.to_string() if len(reason_counts) else "(none)")
    summary_lines.append("")

    fig, ax = plt.subplots(figsize=(9, 5))
    bars = ax.bar(decline_by_device.index, decline_by_device["decline_pct"], color=DECLINE_COLOR)
    ax.set_title("Decline Rate by Device")
    ax.set_ylabel("Declined %")
    ax.bar_label(bars, fmt="%.2f%%", padding=3)
    ax.tick_params(axis='x', rotation=15)
    finish_chart(fig, ax)
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.96, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "07_decline_rate_by_device.png"))
    plt.close()
    gc.collect()

    if len(reason_counts):
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.barh(reason_counts.index[::-1], reason_counts.values[::-1], color=DECLINE_COLOR)
        ax.set_title("Decline Reasons (All Devices)")
        ax.set_xlabel("Number of Declined Transactions")
        finish_chart(fig, ax)
        plt.subplots_adjust(left=0.32, right=0.97, bottom=0.02, top=0.96, wspace=0.3, hspace=0.45)
        plt.savefig(os.path.join(outdir, "08_decline_reasons.png"))
        plt.close()
        gc.collect()

    # ---------- 3. BUSIEST TIME OF DAY (overall, all attempts incl. declines) ----------
    hourly_all = _hourly_series(deposits)
    hourly_declined_all = _hourly_series(deposits[deposits["declined"]])
    busiest_hours = hourly_all.sort_values(ascending=False).head(5)
    summary_lines.append("=== TOP 5 BUSIEST HOURS (overall, all devices, all attempts) ===")
    summary_lines.append(busiest_hours.to_string())
    summary_lines.append("")

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(hourly_all.index, hourly_all.values, color=ACCENT_COLOR, label="Total attempts")
    ax.bar(hourly_declined_all.index, hourly_declined_all.values, color=DECLINE_COLOR, label="Declined")
    ax.set_xticks(range(24))
    ax.set_xlabel("Hour of Day (24h)")
    ax.set_ylabel("Number of Deposits")
    ax.set_title("Overall Busiest Times of Day")
    ax.legend()
    finish_chart(fig, ax, subtitle="All devices combined · declined transactions highlighted in red")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "02_busiest_hours_overall.png"))
    plt.close()
    gc.collect()

    # ---------- 4. HOURLY LOAD CURVE PER DEVICE (comparison) ----------
    hourly_by_device = deposits.pivot_table(
        index="hour", columns="Device Name", values="Amount", aggfunc="count", fill_value=0
    ).reindex(range(24), fill_value=0)

    fig, ax = plt.subplots(figsize=(13, 6))
    for d in devices:
        ax.plot(hourly_by_device.index, hourly_by_device[d], marker="o", markersize=4, label=d, color=colors[d], linewidth=2)
    ax.set_xticks(range(24))
    ax.set_xlabel("Hour of Day (24h)")
    ax.set_ylabel("Number of Deposits")
    ax.set_title("Hourly Usage Pattern by Device")
    ax.legend()
    finish_chart(fig, ax, subtitle="Peak and ramp-up behavior · all attempts (successful + declined)")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "03_hourly_pattern_by_device.png"))
    plt.close()
    gc.collect()

    # ---------- 5. DAY OF WEEK PATTERN (overall) ----------
    dow_counts = deposits["dow"].value_counts().reindex(DOW_ORDER, fill_value=0)
    dow_amount = successful.groupby("dow")["Amount"].sum().reindex(DOW_ORDER, fill_value=0)
    summary_lines.append("=== DAY-OF-WEEK PATTERN (overall) ===")
    summary_lines.append(pd.DataFrame({"transactions": dow_counts, "successful_value": dow_amount}).to_string())
    summary_lines.append("")

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    axes[0].bar(dow_counts.index, dow_counts.values, color=ACCENT_COLOR)
    axes[0].set_title("Transactions by Day of Week")
    axes[0].set_ylabel("Number of Deposits")
    axes[0].tick_params(axis='x', rotation=30)
    bars = axes[1].bar(dow_amount.index, dow_amount.values, color=DEVICE_COLORS[1])
    axes[1].set_title("Successful Value by Day of Week")
    axes[1].set_ylabel("PKR")
    axes[1].bar_label(bars, labels=[fmt_pkr(v) for v in dow_amount.values], padding=3, fontsize=8)
    axes[1].tick_params(axis='x', rotation=30)
    finish_chart(fig, axes, subtitle="All attempts shown on the left; value totals reflect successful deposits only")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "04_day_of_week_pattern.png"))
    plt.close()
    gc.collect()

    # ---------- 6. HEATMAP: Hour x Day-of-week (overall) ----------
    heat = deposits.pivot_table(index="dow", columns="hour", values="Amount", aggfunc="count", fill_value=0)
    heat = heat.reindex(DOW_ORDER).reindex(columns=range(24), fill_value=0)

    fig, ax = plt.subplots(figsize=(14, 5))
    im = ax.imshow(heat.values, aspect="auto", cmap="YlOrRd")
    ax.set_xticks(range(24)); ax.set_xticklabels(range(24))
    ax.set_yticks(range(len(DOW_ORDER))); ax.set_yticklabels(DOW_ORDER)
    ax.set_xlabel("Hour of Day")
    ax.set_title("Usage Heatmap: Day of Week vs Hour")
    for spine in ax.spines.values():
        spine.set_visible(False)
    fig.colorbar(im, ax=ax, label="Number of Deposits")
    fig.text(0.5, 0.955, "All devices combined · all attempts", ha="center", fontsize=9.5, color="#666666", style="italic")
    fig.text(0.99, 0.01, f"Generated {pd.Timestamp.now().strftime('%d %b %Y')} · CDM Usage Analysis", ha="right", va="bottom", fontsize=7.5, color="#999999")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.02, top=0.94, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "05_heatmap_dow_hour.png"))
    plt.close()
    gc.collect()

    # ---------- 7. Per-device heatmaps (combined view) ----------
    fig, axes = plt.subplots(len(devices), 1, figsize=(12, 3.3 * len(devices)))
    if len(devices) == 1:
        axes = [axes]
    for ax, d in zip(axes, devices):
        sub = deposits[deposits["Device Name"] == d]
        h = sub.pivot_table(index="dow", columns="hour", values="Amount", aggfunc="count", fill_value=0)
        h = h.reindex(DOW_ORDER).reindex(columns=range(24), fill_value=0)
        im = ax.imshow(h.values, aspect="auto", cmap="YlOrRd")
        ax.set_xticks(range(24)); ax.set_xticklabels(range(24), fontsize=8)
        ax.set_yticks(range(len(DOW_ORDER))); ax.set_yticklabels(DOW_ORDER, fontsize=8)
        ax.set_title(d, fontsize=12, fontweight="bold")
        for spine in ax.spines.values():
            spine.set_visible(False)
        fig.colorbar(im, ax=ax)
    fig.text(0.99, 0.005, f"Generated {pd.Timestamp.now().strftime('%d %b %Y')} · CDM Usage Analysis", ha="right", va="bottom", fontsize=7.5, color="#999999")
    plt.subplots_adjust(left=0.08, right=0.97, bottom=0.015, top=1, wspace=0.3, hspace=0.45)
    plt.savefig(os.path.join(outdir, "06_heatmap_per_device.png"))
    plt.close()
    gc.collect()

    # ---------- 8. INDIVIDUAL PER-DEVICE REPORTS ----------
    summary_lines.append("=== INDIVIDUAL DEVICE REPORTS ===")
    for i, d in enumerate(devices, start=1):
        sub = deposits[deposits["Device Name"] == d]
        lines, _ = _device_report(d, sub, colors, outdir, i)
        summary_lines.extend(lines)

    # ---------- Write summary text ----------
    summary_text = "\n".join(summary_lines)
    with open(os.path.join(outdir, "summary.txt"), "w") as f:
        f.write(summary_text)

    # ---------- Excel report ----------
    excel_path = os.path.join(outdir, "CDM_Analysis_Report.xlsx")
    export_cols = ["Device Name", "datetime", "hour", "dow", "Amount", "Status",
                    "Result Code", "Result Message", "declined", "decline_reason"]
    all_txns = deposits[export_cols].sort_values("datetime").reset_index(drop=True)

    overview_df = pd.DataFrame({
        "step_name": event_breakdown.index,
        "raw_export_name": event_breakdown.index,
        "row_count": event_breakdown.values,
    })
    overview_df["step_name"] = overview_df["raw_export_name"].map(friendly_step)
    overview_df["is_counted_as_deposit"] = overview_df["raw_export_name"] == "Cash Deposit"
    overview_df = overview_df[["step_name", "row_count", "is_counted_as_deposit", "raw_export_name"]]
    reconciliation = pd.DataFrame({
        "metric": [
            "Total rows in raw file",
            "Cash Deposit rows (= actual deposit attempts)",
            "  of which Successful",
            "  of which Declined",
            "Other session-step rows (not counted as deposits)",
        ],
        "value": [
            total_raw_rows, n_cash_deposit, total_all - total_declined, total_declined, n_other_steps,
        ],
    })

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        reconciliation.to_excel(writer, sheet_name="Data Overview", index=False, startrow=0)
        overview_df.to_excel(writer, sheet_name="Data Overview", index=False, startrow=len(reconciliation) + 2)
        usage.to_excel(writer, sheet_name="Device Usage")
        decline_by_device.to_excel(writer, sheet_name="Decline Summary")
        hourly_all.to_frame("total_deposits").join(
            hourly_declined_all.to_frame("declined_deposits")
        ).to_excel(writer, sheet_name="Hourly (overall)")
        hourly_by_device.to_excel(writer, sheet_name="Hourly by Device")
        pd.DataFrame({"transactions": dow_counts, "successful_value": dow_amount}).to_excel(writer, sheet_name="Day of Week")
        all_txns.to_excel(writer, sheet_name="All Transactions", index=False)

        for d in devices:
            sub = all_txns[all_txns["Device Name"] == d]
            sheet_name = ("TXN_" + "".join(c if c.isalnum() else "_" for c in d))[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

    # ---------- Post-process: highlight declined rows RED ----------
    wb = load_workbook(excel_path)
    txn_sheet_names = ["All Transactions"] + [
        ("TXN_" + "".join(c if c.isalnum() else "_" for c in d))[:31] for d in devices
    ]
    declined_col_idx = export_cols.index("declined") + 1  # 1-indexed for openpyxl

    for sheet_name in txn_sheet_names:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            declined_val = row[declined_col_idx - 1].value
            if declined_val is True or str(declined_val).strip().upper() == "TRUE":
                for cell in row:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
        for i, col in enumerate(export_cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)

    wb.save(excel_path)

    # ---------- Separate CLEAN export: Cash Deposit rows only, ready for downstream use ----------
    clean_path = os.path.join(outdir, "CDM_Cleaned_Deposits.xlsx")
    clean_cols = ["Code", "Device Code", "Device Name", "datetime", "hour", "dow",
                  "Amount", "Status", "declined", "decline_reason"]
    clean_cols = [c for c in clean_cols if c in deposits.columns or c in ["hour", "dow", "declined", "decline_reason"]]
    clean_df = deposits[clean_cols].sort_values("datetime").reset_index(drop=True)
    clean_df = clean_df.rename(columns={
        "Code": "Transaction Code",
        "Device Code": "Device Code",
        "Device Name": "Device Name",
        "datetime": "Date & Time",
        "hour": "Hour",
        "dow": "Day of Week",
        "Amount": "Amount (PKR)",
        "Status": "Status",
        "declined": "Declined",
        "decline_reason": "Decline Reason",
    })

    with pd.ExcelWriter(clean_path, engine="openpyxl") as writer:
        clean_df.to_excel(writer, sheet_name="Cash Deposits", index=False)

    wb_clean = load_workbook(clean_path)
    ws_clean = wb_clean["Cash Deposits"]
    for cell in ws_clean[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    declined_col_idx_clean = list(clean_df.columns).index("Declined") + 1
    for row in ws_clean.iter_rows(min_row=2, max_row=ws_clean.max_row):
        if row[declined_col_idx_clean - 1].value is True:
            for cell in row:
                cell.fill = RED_FILL
                cell.font = RED_FONT
    for i, col in enumerate(clean_df.columns, start=1):
        ws_clean.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)
    ws_clean.freeze_panes = "A2"
    wb_clean.save(clean_path)

    print(summary_text)
    print(f"\nAll outputs saved to: {outdir}")
    return summary_text


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python cdm_analysis.py <input_file> [output_folder]")
        sys.exit(1)
    infile = sys.argv[1]
    outdir = sys.argv[2] if len(sys.argv) > 2 else "cdm_analysis_output"
    data = load_data(infile)
    analyze(data, outdir)
